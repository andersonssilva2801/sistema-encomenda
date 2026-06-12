from datetime import date
from io import BytesIO
from flask import Blueprint, render_template, request, send_file, flash
from flask_login import login_required
from sqlalchemy import func
from app.extensions import db
from app.models.encomenda import Encomenda
from app.models.marketplace import Marketplace
from app.models.transportadora import Transportadora
from app.models.funcionario import Funcionario
from app.forms.consultas import ConsultaForm
from app.utils.helpers import get_periodo_mes_atual

consultas_bp = Blueprint('consultas', __name__, url_prefix='/consultas')


def _build_query(form_data):
    """Constrói a query de encomendas com os filtros aplicados."""
    query = Encomenda.query

    if form_data.get('data_inicio'):
        query = query.filter(Encomenda.data_envio >= form_data['data_inicio'])
    if form_data.get('data_fim'):
        query = query.filter(Encomenda.data_envio <= form_data['data_fim'])
    if form_data.get('marketplace_id'):
        query = query.filter(Encomenda.marketplace_id == form_data['marketplace_id'])
    if form_data.get('transportadora_id'):
        query = query.filter(Encomenda.transportadora_id == form_data['transportadora_id'])
    if form_data.get('funcionario_id'):
        query = query.filter(Encomenda.funcionario_id == form_data['funcionario_id'])
    if form_data.get('tipo_movimento') and form_data['tipo_movimento'] != '0':
        query = query.filter(Encomenda.tipo_movimento == form_data['tipo_movimento'])

    return query.order_by(Encomenda.data_envio.desc(), Encomenda.criado_em.desc())


def _popular_choices_consulta(form):
    """Popula os SelectFields do formulário de consulta."""
    opcao_todos = [(0, 'Todos')]
    form.marketplace_id.choices = opcao_todos + [
        (m.id, m.nome) for m in Marketplace.query.order_by(Marketplace.nome).all()
    ]
    form.transportadora_id.choices = opcao_todos + [
        (t.id, t.nome) for t in Transportadora.query.order_by(Transportadora.nome).all()
    ]
    form.funcionario_id.choices = opcao_todos + [
        (f.id, f.nome) for f in Funcionario.query.order_by(Funcionario.nome).all()
    ]


@consultas_bp.route('/', methods=['GET'])
@login_required
def index():
    form = ConsultaForm(request.args)
    _popular_choices_consulta(form)

    primeiro_dia, ultimo_dia = get_periodo_mes_atual()
    pesquisado = request.args.get('submit') is not None or any(
        request.args.get(k) for k in ['data_inicio', 'data_fim', 'marketplace_id', 'transportadora_id', 'funcionario_id', 'tipo_movimento']
    )

    form_data = {}
    encomendas = []
    total_caixas = 0
    total_registros = 0

    if pesquisado:
        form_data = {
            'data_inicio': form.data_inicio.data,
            'data_fim': form.data_fim.data,
            'marketplace_id': form.marketplace_id.data if form.marketplace_id.data else None,
            'transportadora_id': form.transportadora_id.data if form.transportadora_id.data else None,
            'funcionario_id': form.funcionario_id.data if form.funcionario_id.data else None,
            'tipo_movimento': str(form.tipo_movimento.data) if form.tipo_movimento.data else None,
        }
        encomendas = _build_query(form_data).all()
        total_caixas = sum(e.quantidade_caixas for e in encomendas)
        total_registros = len(encomendas)

    return render_template(
        'consultas/index.html',
        form=form,
        encomendas=encomendas,
        total_caixas=total_caixas,
        total_registros=total_registros,
        pesquisado=pesquisado,
    )


@consultas_bp.route('/exportar-excel')
@login_required
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Biblioteca openpyxl não disponível.', 'danger')
        return render_template('consultas/index.html')

    form_data = {
        'data_inicio': request.args.get('data_inicio') or None,
        'data_fim': request.args.get('data_fim') or None,
        'marketplace_id': request.args.get('marketplace_id', type=int) or None,
        'transportadora_id': request.args.get('transportadora_id', type=int) or None,
        'funcionario_id': request.args.get('funcionario_id', type=int) or None,
        'tipo_movimento': request.args.get('tipo_movimento') or None,
    }

    # Converte strings de data para objetos date
    from datetime import datetime
    for campo in ('data_inicio', 'data_fim'):
        if form_data[campo] and isinstance(form_data[campo], str):
            try:
                form_data[campo] = datetime.strptime(form_data[campo], '%Y-%m-%d').date()
            except ValueError:
                form_data[campo] = None

    encomendas = _build_query(form_data).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Encomendas'

    cabecalho_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    cabecalho_font = Font(color='FFFFFF', bold=True)

    cabecalhos = ['#', 'Data', 'Tipo', 'Marketplace', 'Transportadora', 'Funcionário', 'Qtd Caixas', 'Observações', 'Registrado por']
    for col, titulo in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = cabecalho_fill
        cell.font = cabecalho_font
        cell.alignment = Alignment(horizontal='center')

    for row, enc in enumerate(encomendas, start=2):
        ws.cell(row=row, column=1, value=enc.codigo_formatado)
        ws.cell(row=row, column=2, value=enc.data_envio_formatada)
        ws.cell(row=row, column=3, value=enc.tipo_movimento)
        ws.cell(row=row, column=4, value=enc.marketplace.nome if enc.marketplace else '')
        ws.cell(row=row, column=5, value=enc.transportadora.nome if enc.transportadora else '')
        ws.cell(row=row, column=6, value=enc.funcionario.nome if enc.funcionario else '')
        ws.cell(row=row, column=7, value=enc.quantidade_caixas)
        ws.cell(row=row, column=8, value=enc.observacoes or '')
        ws.cell(row=row, column=9, value=enc.usuario.nome if enc.usuario else '')

    # Linha de totais
    ultima_linha = len(encomendas) + 2
    ws.cell(row=ultima_linha, column=6, value='TOTAL:').font = Font(bold=True)
    ws.cell(row=ultima_linha, column=7, value=sum(e.quantidade_caixas for e in encomendas)).font = Font(bold=True)

    for col in range(1, len(cabecalhos) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].auto_size = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='encomendas.xlsx'
    )
